from datetime import datetime
from io import BytesIO
from pathlib import Path

from fpdf import FPDF, XPos, YPos
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.config import settings

LOGO_PATH = Path(__file__).resolve().parent / "static" / "logo-muni.png"
LOGO_RATIO = 241 / 286


def _latin(text: str | None) -> str:
    if text is None:
        return ""
    return str(text).encode("latin-1", errors="replace").decode("latin-1")


def _clip(text: str | None, limit: int = 45) -> str:
    value = _latin(text)
    return value if len(value) <= limit else value[: limit - 1] + "..."


def _generated_at() -> str:
    return datetime.now().strftime("%d/%m/%Y %H:%M:%S")


def _excel_sheet(rows: list[list], title: str, report_title: str) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = title

    headers = rows[0]
    ncols = len(headers)

    ws.append([_latin(settings.entity_name)])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(row=1, column=1).font = Font(bold=True, size=13, color="1D64F1")

    ws.append([f"{report_title}  |  Generado el {_generated_at()}"])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws.cell(row=2, column=1).font = Font(size=9, color="666666")

    ws.append([])
    header_row = 4
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1D64F1")
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, ncols + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in rows[1:]:
        ws.append(row)

    for col in range(1, ncols + 1):
        width = max(len(str(r[col - 1])) for r in rows) + 3
        ws.column_dimensions[get_column_letter(col)].width = max(width, 12)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def _pdf_header(pdf: FPDF, title: str):
    logo_h = 15
    logo_w = logo_h * LOGO_RATIO
    if LOGO_PATH.exists():
        pdf.image(
            str(LOGO_PATH),
            x=pdf.w - pdf.r_margin - logo_w,
            y=pdf.t_margin,
            w=logo_w,
        )
    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 7, _latin(settings.entity_name), new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 6, "Sistema de Almacen e Inventario", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
    pdf.ln(1)
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, title, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
    pdf.set_font("Helvetica", "", 8)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 6, f"Generado el {_generated_at()}", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="R")
    pdf.set_text_color(0, 0, 0)
    pdf.ln(2)


def _pdf_table(title: str, headers: list[str], widths: list[int], data: list[list]) -> BytesIO:
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)

    def draw_header():
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_fill_color(29, 100, 241)
        pdf.set_text_color(255, 255, 255)
        for i, header in enumerate(headers):
            pdf.cell(widths[i], 8, _latin(header), border=1, align="C", fill=True)
        pdf.ln()
        pdf.set_text_color(0, 0, 0)

    pdf.add_page()
    _pdf_header(pdf, title)

    draw_header()
    pdf.set_font("Helvetica", "", 8)
    pdf.set_fill_color(238, 246, 255)
    fill = False
    for row in data:
        if pdf.get_y() > 260:
            pdf.add_page()
            draw_header()
            pdf.set_font("Helvetica", "", 8)
            pdf.set_fill_color(238, 246, 255)
        for i, value in enumerate(row):
            align = "C" if i not in (1, 3) else "L"
            pdf.cell(widths[i], 8, _clip(value, 60), border=1, align=align, fill=fill)
        pdf.ln()
        fill = not fill

    return BytesIO(bytes(pdf.output(dest="S")))


def products_excel(items) -> BytesIO:
    rows = [["Codigo", "Nombre", "Stock actual", "Creado"]]
    for p in items:
        rows.append([p.code, p.name, p.current_stock, p.created_at.date().isoformat()])
    return _excel_sheet(rows, "Productos", "Reporte de productos")


def products_pdf(items) -> BytesIO:
    headers = ["Codigo", "Nombre", "Stock", "Creado"]
    widths = [40, 90, 20, 30]
    data = [
        [p.code, p.name, p.current_stock, p.created_at.date().isoformat()] for p in items
    ]
    return _pdf_table("Lista de productos", headers, widths, data)


def movements_excel(items) -> BytesIO:
    rows = [
        ["Fecha", "Producto", "Codigo", "Tipo", "Cantidad", "Stock resultante", "Nota"]
    ]
    for m in items:
        rows.append(
            [
                m.movement_date.isoformat(),
                m.product.name,
                m.product.code,
                "Entrada" if m.movement_type == "ENTRADA" else "Salida",
                m.quantity,
                m.stock_after,
                m.note or "",
            ]
        )
    return _excel_sheet(rows, "Movimientos", "Reporte de movimientos")


def movements_pdf(items) -> BytesIO:
    headers = ["Fecha", "Producto", "Tipo", "Cant.", "Stock", "Nota"]
    widths = [24, 72, 20, 16, 16, 32]
    data = []
    for m in items:
        data.append(
            [
                m.movement_date.isoformat(),
                f"{m.product.code} - {m.product.name}",
                "Entrada" if m.movement_type == "ENTRADA" else "Salida",
                m.quantity,
                m.stock_after,
                m.note or "",
            ]
        )
    return _pdf_table("Lista de movimientos", headers, widths, data)